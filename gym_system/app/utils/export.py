"""
Export utilities for generating Excel and PDF reports.
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def export_to_excel(data, headers, title="Report"):
    """
    Export data to Excel file.

    Args:
        data: List of dicts or list of lists
        headers: Dict mapping keys to Arabic labels or list of labels
        title: Sheet title

    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Determine column keys and labels
    if isinstance(headers, dict):
        col_keys = list(headers.keys())
        col_labels = list(headers.values())
    else:
        col_keys = None
        col_labels = headers

    # Write headers
    for col_num, label in enumerate(col_labels, 1):
        cell = ws.cell(row=1, column=col_num, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = max(len(str(label)) * 2, 15)

    # Write data
    for row_num, row_data in enumerate(data, 2):
        if col_keys and isinstance(row_data, dict):
            values = [row_data.get(key, '') for key in col_keys]
        elif isinstance(row_data, dict):
            values = list(row_data.values())
        else:
            values = row_data

        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # RTL direction for Arabic
    ws.sheet_view.rightToLeft = True

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def export_members_excel(members):
    """Export members list to Excel"""
    headers = {
        'id': '#',
        'name': 'الاسم',
        'phone': 'الهاتف',
        'email': 'البريد',
        'brand': 'البراند',
        'status': 'الحالة',
        'subscription': 'الاشتراك',
        'created_at': 'تاريخ التسجيل'
    }

    data = []
    for member in members:
        status_map = {
            'active': 'نشط',
            'expired': 'منتهي',
            'frozen': 'مجمد',
            'no_subscription': 'بدون اشتراك'
        }

        data.append({
            'id': member.id,
            'name': member.name,
            'phone': member.phone,
            'email': member.email or '-',
            'brand': member.brand.name,
            'status': status_map.get(member.subscription_status, '-'),
            'subscription': member.active_subscription.plan.name if member.active_subscription else '-',
            'created_at': member.created_at.strftime('%Y-%m-%d')
        })

    return export_to_excel(data, headers, 'الأعضاء')


def export_attendance_excel(attendances):
    """Export attendance records to Excel"""
    headers = {
        'date': 'التاريخ',
        'time': 'الوقت',
        'member_name': 'العضو',
        'phone': 'الهاتف',
        'brand': 'البراند',
        'source': 'المصدر'
    }

    source_map = {
        'manual': 'يدوي',
        'fingerprint': 'بصمة',
        'qr': 'QR'
    }

    data = []
    for att in attendances:
        data.append({
            'date': att.check_in.strftime('%Y-%m-%d'),
            'time': att.check_in.strftime('%H:%M:%S'),
            'member_name': att.member.name,
            'phone': att.member.phone,
            'brand': att.brand.name,
            'source': source_map.get(att.source, att.source)
        })

    return export_to_excel(data, headers, 'الحضور')


def export_financial_excel(report_data):
    """Export financial report to Excel"""
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = 'الملخص'
    ws_summary.sheet_view.rightToLeft = True

    summary_data = [
        ['إجمالي الدخل', report_data['total_income']],
        ['إجمالي المصروفات', report_data['total_expenses']],
        ['إجمالي الرواتب', report_data['total_salaries']],
        ['صافي الربح', report_data['net_profit']]
    ]

    for row_num, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_num, column=1, value=label)
        ws_summary.cell(row=row_num, column=2, value=f"{value:,.0f} ر.س")

    # Daily breakdown sheet
    ws_daily = wb.create_sheet('التفاصيل اليومية')
    ws_daily.sheet_view.rightToLeft = True

    headers = ['التاريخ', 'الدخل', 'المصروفات', 'الصافي']
    for col_num, header in enumerate(headers, 1):
        cell = ws_daily.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for row_num, day in enumerate(report_data.get('daily_breakdown', []), 2):
        ws_daily.cell(row=row_num, column=1, value=day['date'])
        ws_daily.cell(row=row_num, column=2, value=day['income'])
        ws_daily.cell(row=row_num, column=3, value=day['expenses'])
        ws_daily.cell(row=row_num, column=4, value=day['net'])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def generate_filename(prefix, extension='xlsx'):
    """Generate a filename with timestamp"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f"{prefix}_{timestamp}.{extension}"
