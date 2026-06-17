from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
import io
from flask_wtf import FlaskForm
from wtforms import StringField, DecimalField, TextAreaField, DateField, SelectField
from wtforms.validators import DataRequired, Optional
from datetime import date, datetime

from app import db
from app.models.company import Brand
from app.models.user import User
from app.models.finance import Income, Expense, Salary, Refund, ExpenseCategory
from app.models.daily_closing import DailyClosing
from app.models.subscription import Subscription, SubscriptionPayment
from app.utils.decorators import finance_required, brand_manager_required
from app.utils.helpers import pagination_args, save_uploaded_file, apply_branch_filter, check_entity_access

finance_bp = Blueprint('finance', __name__)


@finance_bp.route('/dashboard')
@login_required
@finance_required
def dashboard():
    """Finance dashboard with daily summary"""
    today = date.today()

    # Base query for brand filtering
    if current_user.can_view_all_brands:
        brand_id = request.args.get('brand_id', type=int)
    else:
        brand_id = current_user.brand_id

    # Calculate today's total sales (GYM-32: skip soft-deleted)
    sales_query = db.session.query(db.func.sum(Income.amount)).filter(
        Income.date == today,
        Income.is_deleted == False,
    )
    if brand_id:
        sales_query = sales_query.filter(Income.brand_id == brand_id)
    total_sales = sales_query.scalar() or 0

    # Calculate today's total expenses (GYM-32: skip soft-deleted)
    expenses_query = db.session.query(db.func.sum(Expense.amount)).filter(
        Expense.date == today,
        Expense.is_deleted == False,
    )
    if brand_id:
        expenses_query = expenses_query.filter(Expense.brand_id == brand_id)
    total_expenses = expenses_query.scalar() or 0

    # Get today's daily closing (if exists)
    closing_query = DailyClosing.query.filter(
        DailyClosing.closing_date == today
    )
    if brand_id:
        closing_query = closing_query.filter(DailyClosing.brand_id == brand_id)
    daily_closing = closing_query.first()

    # Calculate cash difference
    cash_difference = None
    if daily_closing:
        expected_cash = float(daily_closing.cash_amount or 0)
        actual_cash = float(daily_closing.actual_cash_submitted or 0)
        cash_difference = actual_cash - expected_cash

    # Financial alerts
    alerts = []

    # Alert: Day not closed yet
    if not daily_closing and datetime.now().hour >= 22:
        alerts.append({
            'type': 'warning',
            'message': 'لم يتم إقفال اليوم بعد'
        })

    # Alert: Large cash difference
    if cash_difference and abs(cash_difference) > 100:
        alerts.append({
            'type': 'danger',
            'message': f'فرق كبير في الكاش: {cash_difference:,.2f} ر.س'
        })

    # Alert: No sales today
    if total_sales == 0 and datetime.now().hour >= 12:
        alerts.append({
            'type': 'info',
            'message': 'لا توجد مبيعات اليوم حتى الآن'
        })

    # Get recent payments (last 10 today). GYM-32 — drop payments tied to
    # soft-deleted subscriptions; otherwise refunds/cancellations would
    # ghost-appear in "آخر المبيعات اليوم".
    payments_query = SubscriptionPayment.query.join(Subscription).filter(
        db.func.date(SubscriptionPayment.payment_date) == today,
        Subscription.is_deleted == False,
    )
    if brand_id:
        payments_query = payments_query.filter(Subscription.brand_id == brand_id)
    recent_payments = payments_query.order_by(
        SubscriptionPayment.payment_date.desc()
    ).limit(10).all()

    # Get recent expenses (last 5 today; GYM-32 — skip soft-deleted)
    expenses_list_query = Expense.query.filter(
        Expense.date == today,
        Expense.is_deleted == False,
    )
    if brand_id:
        expenses_list_query = expenses_list_query.filter(Expense.brand_id == brand_id)
    recent_expenses = expenses_list_query.order_by(
        Expense.created_at.desc()
    ).limit(5).all()

    # Get pending expenses count
    pending_expenses_query = Expense.query.filter_by(status='pending').filter(
        Expense.is_deleted == False  # GYM-32
    )
    if brand_id:
        pending_expenses_query = pending_expenses_query.filter(Expense.brand_id == brand_id)
    pending_expenses_count = pending_expenses_query.count()

    # Get brands for filter
    brands = None
    if current_user.can_view_all_brands:
        brands = Brand.query.filter_by(is_active=True).all()

    return render_template('finance/dashboard.html',
                          total_sales=total_sales,
                          total_expenses=total_expenses,
                          cash_difference=cash_difference,
                          alerts=alerts,
                          recent_payments=recent_payments,
                          recent_expenses=recent_expenses,
                          daily_closing=daily_closing,
                          brands=brands,
                          pending_expenses_count=pending_expenses_count,
                          today=today)


class ExpenseForm(FlaskForm):
    """Expense form"""
    category_name = SelectField('الفئة', validators=[DataRequired()])
    amount = DecimalField('المبلغ', validators=[DataRequired()])
    description = TextAreaField('الوصف')
    date = DateField('التاريخ', default=date.today, validators=[DataRequired()])


class SalaryForm(FlaskForm):
    """Salary form"""
    base_salary = DecimalField('الراتب الأساسي', validators=[DataRequired()])
    deductions = DecimalField('الخصومات', default=0)
    bonuses = DecimalField('البدلات', default=0)
    notes = TextAreaField('ملاحظات')


@finance_bp.route('/sales-transactions')
@login_required
@finance_required
def sales_transactions():
    """Detailed sales transactions view"""
    page, per_page = pagination_args(request)
    date_from = request.args.get('date_from', date.today().replace(day=1).isoformat())
    date_to = request.args.get('date_to', date.today().isoformat())

    # Parse dates
    try:
        from_date = datetime.fromisoformat(date_from).date()
        to_date = datetime.fromisoformat(date_to).date()
    except:
        from_date = date.today().replace(day=1)
        to_date = date.today()

    # Base query - join with subscription and member to get all details
    query = db.session.query(SubscriptionPayment).join(Subscription).join(
        Subscription.member
    )

    # Brand filter
    if current_user.can_view_all_brands:
        brand_id = request.args.get('brand_id', type=int)
        if brand_id:
            query = query.filter(Subscription.brand_id == brand_id)
    else:
        query = query.filter(Subscription.brand_id == current_user.brand_id)

    # Date filter (GYM-32 — skip soft-deleted subscriptions)
    query = query.filter(
        db.func.date(SubscriptionPayment.payment_date) >= from_date,
        db.func.date(SubscriptionPayment.payment_date) <= to_date,
        Subscription.is_deleted == False,
    )

    # Payment method filter
    payment_method = request.args.get('payment_method', '')
    if payment_method:
        query = query.filter(SubscriptionPayment.payment_method == payment_method)

    # Service filter (subscription plan)
    service_id = request.args.get('service_id', type=int)
    if service_id:
        query = query.filter(Subscription.plan_id == service_id)

    # Receptionist filter
    receptionist_id = request.args.get('receptionist_id', type=int)
    if receptionist_id:
        query = query.filter(Subscription.created_by == receptionist_id)

    # Member search
    search = request.args.get('search', '').strip()
    if search:
        from app.models.member import Member
        query = query.filter(Member.name.like(f'%{search}%'))

    # Calculate totals
    total_amount = query.with_entities(
        db.func.sum(SubscriptionPayment.amount)
    ).scalar() or 0

    # Count by payment method
    payment_method_counts = {}
    if current_user.can_view_all_brands and not brand_id:
        base_query = db.session.query(SubscriptionPayment).join(Subscription)
    else:
        base_query = query

    for method in ['cash', 'card', 'transfer']:
        count_query = base_query.filter(
            SubscriptionPayment.payment_method == method,
            db.func.date(SubscriptionPayment.payment_date) >= from_date,
            db.func.date(SubscriptionPayment.payment_date) <= to_date
        )
        if not current_user.can_view_all_brands or brand_id:
            payment_method_counts[method] = count_query.count()
        else:
            payment_method_counts[method] = count_query.count()

    # Pagination
    transactions = query.order_by(SubscriptionPayment.payment_date.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    # Get filter options
    brands = None
    if current_user.can_view_all_brands:
        brands = Brand.query.filter_by(is_active=True).all()

    # Get subscription plans for service filter
    from app.models.subscription import Plan
    if current_user.can_view_all_brands:
        if brand_id:
            services = Plan.query.filter_by(
                brand_id=brand_id, is_active=True
            ).all()
        else:
            services = Plan.query.filter_by(is_active=True).all()
    else:
        services = Plan.query.filter_by(
            brand_id=current_user.brand_id, is_active=True
        ).all()

    # Get receptionists for filter
    receptionists = User.query.filter(
        User.role.has(name_en='receptionist')
    )
    if not current_user.can_view_all_brands:
        receptionists = receptionists.filter_by(brand_id=current_user.brand_id)
    receptionists = receptionists.all()

    return render_template('finance/sales_transactions.html',
                          transactions=transactions,
                          brands=brands,
                          services=services,
                          receptionists=receptionists,
                          total_amount=total_amount,
                          payment_method_counts=payment_method_counts,
                          date_from=date_from,
                          date_to=date_to,
                          payment_method=payment_method,
                          service_id=service_id,
                          receptionist_id=receptionist_id,
                          search=search)


@finance_bp.route('/income')
@login_required
@finance_required
def income_list():
    """List income records"""
    page, per_page = pagination_args(request)
    date_from = request.args.get('date_from', date.today().replace(day=1).isoformat())
    date_to = request.args.get('date_to', date.today().isoformat())
    payment_method = request.args.get('payment_method', '')

    # Parse dates
    try:
        from_date = date.fromisoformat(date_from)
        to_date = date.fromisoformat(date_to)
    except:
        from_date = date.today().replace(day=1)
        to_date = date.today()

    # Base query
    if current_user.can_view_all_brands:
        brand_id = request.args.get('brand_id', type=int)
        if brand_id:
            query = Income.query.filter_by(brand_id=brand_id)
        else:
            query = Income.query
    elif current_user.branch_id:
        # Branch accountant - only see their branch
        query = Income.query.filter_by(brand_id=current_user.brand_id, branch_id=current_user.branch_id)
    else:
        # Central accountant - see all branches in brand
        query = Income.query.filter_by(brand_id=current_user.brand_id)

    # Date filter (GYM-32 — skip soft-deleted)
    query = query.filter(Income.date >= from_date, Income.date <= to_date,
                         Income.is_deleted == False)

    # Payment method filter
    if payment_method:
        query = query.filter(Income.payment_method == payment_method)

    # Calculate totals by payment method
    base_filter = [Income.date >= from_date, Income.date <= to_date,
                   Income.is_deleted == False]
    if current_user.brand_id:
        base_filter.append(Income.brand_id == current_user.brand_id)
        # Branch accountant filter
        if current_user.branch_id:
            base_filter.append(Income.branch_id == current_user.branch_id)

    total = db.session.query(db.func.sum(Income.amount)).filter(*base_filter).scalar() or 0

    # Payment breakdown
    payment_breakdown = db.session.query(
        Income.payment_method,
        db.func.sum(Income.amount)
    ).filter(*base_filter).group_by(Income.payment_method).all()

    payment_stats = {'cash': 0, 'card': 0, 'transfer': 0}
    for method, amount in payment_breakdown:
        if method in payment_stats:
            payment_stats[method] = float(amount or 0)

    # Pagination
    income = query.order_by(Income.date.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    # Brands for filter
    brands = None
    if current_user.can_view_all_brands:
        brands = Brand.query.filter_by(is_active=True).all()

    return render_template('finance/income.html',
                          income=income,
                          brands=brands,
                          total=total,
                          payment_stats=payment_stats,
                          payment_method=payment_method,
                          date_from=date_from,
                          date_to=date_to)


@finance_bp.route('/expenses')
@login_required
@finance_required
def expenses():
    """List expense records"""
    page, per_page = pagination_args(request)
    date_from = request.args.get('date_from', date.today().replace(day=1).isoformat())
    date_to = request.args.get('date_to', date.today().isoformat())
    status = request.args.get('status', '')
    category = request.args.get('category', '')
    user_id = request.args.get('user_id', type=int)
    branch_id_filter = request.args.get('branch_id', type=int)

    try:
        from_date = date.fromisoformat(date_from)
        to_date = date.fromisoformat(date_to)
    except:
        from_date = date.today().replace(day=1)
        to_date = date.today()

    # Base query
    if current_user.can_view_all_brands:
        brand_id = request.args.get('brand_id', type=int)
        if brand_id:
            query = Expense.query.filter_by(brand_id=brand_id)
        else:
            query = Expense.query
    elif current_user.branch_id:
        # Branch accountant - only see their branch
        query = Expense.query.filter_by(brand_id=current_user.brand_id, branch_id=current_user.branch_id)
    else:
        # Central accountant - see all branches in brand
        query = Expense.query.filter_by(brand_id=current_user.brand_id)

    query = query.filter(Expense.date >= from_date, Expense.date <= to_date)
    # GYM-32 — hide soft-deleted expenses
    query = query.filter(Expense.is_deleted == False)

    # Additional filters
    if status:
        query = query.filter(Expense.status == status)

    if category:
        query = query.filter(Expense.category_name == category)

    if user_id:
        query = query.filter(Expense.created_by == user_id)

    # Branch filter for central accountants
    if branch_id_filter and not current_user.branch_id:
        query = query.filter(Expense.branch_id == branch_id_filter)

    # Calculate totals
    base_filter = [
        Expense.date >= from_date, Expense.date <= to_date,
        Expense.is_deleted == False,
    ]
    if current_user.brand_id:
        base_filter.append(Expense.brand_id == current_user.brand_id)
        # Branch accountant filter
        if current_user.branch_id:
            base_filter.append(Expense.branch_id == current_user.branch_id)

    total = db.session.query(db.func.sum(Expense.amount)).filter(*base_filter).scalar() or 0
    approved_total = db.session.query(db.func.sum(Expense.amount)).filter(
        *base_filter, Expense.status == 'approved'
    ).scalar() or 0
    pending_count = Expense.query.filter(*base_filter, Expense.status == 'pending').count()

    # Pagination
    expenses = query.order_by(Expense.date.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    # Brands for filter
    brands = None
    if current_user.can_view_all_brands:
        brands = Brand.query.filter_by(is_active=True).all()

    # Branches for filter (central accountants only)
    from app.models.company import Branch
    branches = None
    if current_user.brand_id and not current_user.branch_id:
        branches = Branch.query.filter_by(brand_id=current_user.brand_id, is_active=True).all()

    # Users for filter
    from app.models.user import User
    users = User.query.filter_by(brand_id=current_user.brand_id, is_active=True).all() if current_user.brand_id else []

    # Categories for filter
    categories = ['رواتب', 'إيجار', 'كهرباء', 'ماء', 'صيانة', 'معدات', 'تسويق', 'مستلزمات', 'أخرى']

    return render_template('finance/expenses.html',
                          expenses=expenses,
                          brands=brands,
                          branches=branches,
                          users=users,
                          categories=categories,
                          total=total,
                          approved_total=approved_total,
                          pending_count=pending_count,
                          status=status,
                          category=category,
                          user_id=user_id,
                          branch_id_filter=branch_id_filter,
                          date_from=date_from,
                          date_to=date_to)


@finance_bp.route('/expenses/create', methods=['GET', 'POST'])
@login_required
@finance_required
def expenses_create():
    """Create new expense"""
    form = ExpenseForm()

    # Expense categories
    categories = [
        ('رواتب', 'رواتب'),
        ('إيجار', 'إيجار'),
        ('كهرباء', 'كهرباء'),
        ('ماء', 'ماء'),
        ('صيانة', 'صيانة'),
        ('معدات', 'معدات'),
        ('تسويق', 'تسويق'),
        ('مستلزمات', 'مستلزمات'),
        ('أخرى', 'أخرى'),
    ]
    form.category_name.choices = categories

    # Get brand
    if current_user.is_owner:
        brand_id = request.args.get('brand_id', type=int)
        if not brand_id:
            flash('يرجى اختيار البراند', 'warning')
            return redirect(url_for('admin.brands_list'))
        brand = Brand.query.get_or_404(brand_id)
    else:
        brand = current_user.brand
        brand_id = brand.id

    if form.validate_on_submit():
        # GYM-31 — double-submit dedupe. Key by (brand, branch, category,
        # amount, 30-second window) — receptionist double-clicking save
        # gets a redirect instead of two identical expense rows.
        from datetime import datetime as _dt, timedelta as _td
        _dup = Expense.query.filter(
            Expense.brand_id == brand_id,
            Expense.branch_id == current_user.branch_id,
            Expense.category_name == form.category_name.data,
            Expense.amount == form.amount.data,
            Expense.created_at >= _dt.utcnow() - _td(seconds=30),
            Expense.is_deleted == False,
        ).order_by(Expense.created_at.desc()).first()
        if _dup:
            flash(f'تم تسجيل المصروف للتو (#{_dup.id}) — لم نقم بإنشاء تكرار.', 'info')
            return redirect(url_for('finance.expenses'))

        expense = Expense(
            brand_id=brand_id,
            branch_id=current_user.branch_id,
            category_name=form.category_name.data,
            amount=form.amount.data,
            description=form.description.data,
            date=form.date.data,
            created_by=current_user.id
        )

        # Approval policy: if the creator can already approve expenses,
        # auto-approve. Otherwise the row lands in the pending queue handled
        # by /finance/expenses/pending.
        can_approve = bool(current_user.role and current_user.role.can_approve_expenses)
        if can_approve:
            expense.status = 'approved'
            expense.approved_by = current_user.id
            expense.approved_at = datetime.utcnow()
            flash_msg = 'تم تسجيل المصروف بنجاح'
        else:
            expense.status = 'pending'
            flash_msg = 'تم تسجيل المصروف وهو بانتظار الموافقة'

        # Handle receipt image
        if 'receipt_image' in request.files:
            receipt_file = request.files['receipt_image']
            if receipt_file.filename:
                receipt_path = save_uploaded_file(receipt_file, 'receipts')
                if receipt_path:
                    expense.receipt_image = receipt_path

        db.session.add(expense)
        db.session.commit()

        flash(flash_msg, 'success' if can_approve else 'info')
        return redirect(url_for('finance.expenses'))

    return render_template('finance/expense_form.html', form=form, brand=brand)


@finance_bp.route('/expenses/pending')
@login_required
@finance_required
def pending_expenses():
    """View pending expenses for approval"""
    page, per_page = pagination_args(request)

    # Base query for pending expenses (GYM-32: skip soft-deleted)
    if current_user.can_view_all_brands:
        brand_id = request.args.get('brand_id', type=int)
        if brand_id:
            query = Expense.query.filter_by(brand_id=brand_id, status='pending')
        else:
            query = Expense.query.filter_by(status='pending')
    else:
        query = Expense.query.filter_by(brand_id=current_user.brand_id, status='pending')
    query = query.filter(Expense.is_deleted == False)

    # Pagination
    expenses = query.order_by(Expense.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    # Get brands for filter
    brands = None
    if current_user.can_view_all_brands:
        brands = Brand.query.filter_by(is_active=True).all()

    # Calculate total pending amount (GYM-32: skip soft-deleted)
    total_pending = db.session.query(db.func.sum(Expense.amount)).filter(
        Expense.status == 'pending',
        Expense.is_deleted == False,
    )
    if not current_user.can_view_all_brands:
        total_pending = total_pending.filter(Expense.brand_id == current_user.brand_id)
    total_pending = total_pending.scalar() or 0

    return render_template('finance/pending_expenses.html',
                          expenses=expenses,
                          brands=brands,
                          total_pending=total_pending)


@finance_bp.route('/expenses/<int:expense_id>/approve', methods=['POST'])
@login_required
@finance_required
def approve_expense(expense_id):
    """Approve an expense"""
    expense = Expense.query.get_or_404(expense_id)

    if not current_user.can_access_brand(expense.brand_id):
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('finance.pending_expenses'))

    if expense.status != 'pending':
        flash('هذا المصروف تمت معالجته مسبقاً', 'warning')
        return redirect(url_for('finance.pending_expenses'))

    expense.status = 'approved'
    expense.approved_by = current_user.id
    expense.approved_at = datetime.utcnow()
    db.session.commit()

    flash('تم اعتماد المصروف بنجاح', 'success')
    return redirect(url_for('finance.pending_expenses'))


@finance_bp.route('/expenses/<int:expense_id>/reject', methods=['POST'])
@login_required
@finance_required
def reject_expense(expense_id):
    """Reject an expense"""
    expense = Expense.query.get_or_404(expense_id)

    if not current_user.can_access_brand(expense.brand_id):
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('finance.pending_expenses'))

    if expense.status != 'pending':
        flash('هذا المصروف تمت معالجته مسبقاً', 'warning')
        return redirect(url_for('finance.pending_expenses'))

    reason = request.form.get('reason', '')

    expense.status = 'rejected'
    expense.approved_by = current_user.id
    expense.approved_at = datetime.utcnow()
    expense.rejection_reason = reason
    db.session.commit()

    flash('تم رفض المصروف', 'success')
    return redirect(url_for('finance.pending_expenses'))


@finance_bp.route('/salaries')
@login_required
@finance_required
def salaries_list():
    """List salaries for the period.

    Every active employee with a base salary_amount > 0 in scope appears,
    whether or not a Salary row exists for the month yet. Rows without a
    Salary record are flagged 'غير منشأ' — the owner just clicks
    "إنشاء رواتب الشهر" once and they all materialize as pending Salary rows.
    """
    from app.models.employee import EmployeeReward, EmployeeDeduction

    month = request.args.get('month', date.today().month, type=int)
    year = request.args.get('year', date.today().year, type=int)

    from calendar import monthrange
    period_start = date(year, month, 1)
    period_end = date(year, month, monthrange(year, month)[1])

    # 1) All active employees in scope with a base salary configured.
    users_q = User.query.filter(
        User.is_active == True,
        User.salary_amount != None,
        User.salary_amount > 0,
    )
    if not current_user.can_view_all_brands:
        if current_user.brand_id:
            users_q = users_q.filter(User.brand_id == current_user.brand_id)
        if current_user.branch_id:
            users_q = users_q.filter(
                db.or_(User.branch_id == current_user.branch_id,
                       User.branch_id.is_(None))
            )
    candidates = users_q.order_by(User.name).all()

    # 2) Salary rows for the period, keyed by user_id.
    salary_q = apply_branch_filter(Salary.query, Salary).filter_by(month=month, year=year)
    salary_by_user = {s.user_id: s for s in salary_q.all()}

    # 3) Single batched read for the period's rewards / deductions.
    visible_ids = [u.id for u in candidates] or [-1]
    rewards_q = EmployeeReward.query.filter(
        EmployeeReward.user_id.in_(visible_ids),
        EmployeeReward.is_active == True,
        db.or_(
            db.and_(EmployeeReward.is_recurring == False,
                    EmployeeReward.effective_date >= period_start,
                    EmployeeReward.effective_date <= period_end),
            EmployeeReward.is_recurring == True,
        ),
    ).all()
    deductions_q = EmployeeDeduction.query.filter(
        EmployeeDeduction.user_id.in_(visible_ids),
        EmployeeDeduction.deduction_date >= period_start,
        EmployeeDeduction.deduction_date <= period_end,
    ).all()
    rewards_by_user = {}
    deductions_by_user = {}
    for r in rewards_q:
        rewards_by_user.setdefault(r.user_id, []).append(r)
    for d in deductions_q:
        deductions_by_user.setdefault(d.user_id, []).append(d)

    salary_data = []
    for user in candidates:
        salary = salary_by_user.get(user.id)
        rewards = rewards_by_user.get(user.id, [])
        deductions = deductions_by_user.get(user.id, [])
        total_rewards = sum(float(r.amount) for r in rewards)
        total_deductions = sum(float(d.amount) for d in deductions)
        base = float((salary.base_salary if salary else None) or user.salary_amount or 0)
        net_salary = base + total_rewards - total_deductions

        salary_data.append({
            'user': user,
            'salary': salary,
            'has_record': salary is not None,
            'base_salary': base,
            'rewards': rewards,
            'deductions': deductions,
            'total_rewards': total_rewards,
            'total_deductions': total_deductions,
            'net_salary': net_salary,
        })

    total_base = sum(s['base_salary'] for s in salary_data)
    total_rewards = sum(s['total_rewards'] for s in salary_data)
    total_deductions = sum(s['total_deductions'] for s in salary_data)
    total_net = total_base + total_rewards - total_deductions

    brands = None
    if current_user.can_view_all_brands:
        brands = Brand.query.filter_by(is_active=True).all()

    return render_template('finance/salaries.html',
                          salary_data=salary_data,
                          brands=brands,
                          total_base=total_base,
                          total_rewards=total_rewards,
                          total_deductions=total_deductions,
                          total_net=total_net,
                          month=month,
                          year=year)


@finance_bp.route('/salaries/generate', methods=['POST'])
@login_required
@finance_required
def salaries_generate():
    """One-click salary-row generator for a given month/year.

    For every active user in scope with a base salary_amount > 0, insert a
    Salary row if one doesn't already exist for (user, month, year). Owners
    use this once a month to seed pending salaries that they then 'تأكيد الدفع'
    individually.
    """
    from app.models.user import User
    try:
        month = int(request.form.get('month', date.today().month))
        year = int(request.form.get('year', date.today().year))
    except (TypeError, ValueError):
        month, year = date.today().month, date.today().year

    # Brand scope — admins can pass an explicit brand, everyone else gets theirs.
    if current_user.can_view_all_brands:
        brand_id = request.form.get('brand_id', type=int)
    else:
        brand_id = current_user.brand_id

    users_q = User.query.filter(
        User.is_active == True,
        User.salary_amount != None,
        User.salary_amount > 0,
    )
    if brand_id:
        users_q = users_q.filter(User.brand_id == brand_id)
    candidates = users_q.all()

    inserted = 0
    skipped = 0
    for u in candidates:
        exists = Salary.query.filter_by(user_id=u.id, month=month, year=year).first()
        if exists:
            skipped += 1
            continue
        base = float(u.salary_amount)
        db.session.add(Salary(
            user_id=u.id,
            brand_id=u.brand_id,
            month=month, year=year,
            base_salary=base,
            deductions=0, bonuses=0,
            net_salary=base,
            status='pending',
        ))
        inserted += 1
    db.session.commit()

    if inserted:
        flash(f'تم إنشاء {inserted} راتب جديد لـ {month}/{year} ({skipped} موجود بالفعل)', 'success')
    else:
        flash(f'لا يوجد رواتب جديدة لإنشائها — كل الموظفين لديهم سجل لـ {month}/{year}', 'info')
    return redirect(url_for('finance.salaries_list', month=month, year=year))


@finance_bp.route('/salaries/<int:salary_id>/pay', methods=['POST'])
@login_required
@finance_required
def salary_pay(salary_id):
    """Mark a Salary as paid AND post a matching رواتب expense in one txn.

    GYM-18: salaries previously had a 'paid' status but nothing in the route
    layer ever flipped it. Accountants were creating a manual رواتب expense
    and the employee's salary card stayed perpetually 'pending'. Both happen
    here atomically — the salary row becomes the source of truth for *when*
    a salary was paid, the Expense row keeps the books balanced.
    """
    from app.models.employee import EmployeeReward, EmployeeDeduction
    salary = Salary.query.get_or_404(salary_id)
    if not check_entity_access(salary):
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('finance.salaries_list'))
    if salary.status == 'paid':
        flash('هذا الراتب مدفوع بالفعل', 'info')
        return redirect(url_for('finance.salaries_list', month=salary.month, year=salary.year))

    # Recompute net at pay time (catches rewards/deductions added after the
    # Salary row was first written).
    from calendar import monthrange
    period_start = date(salary.year, salary.month, 1)
    period_end = date(salary.year, salary.month, monthrange(salary.year, salary.month)[1])
    total_rewards = float(db.session.query(db.func.coalesce(db.func.sum(EmployeeReward.amount), 0)).filter(
        EmployeeReward.user_id == salary.user_id,
        EmployeeReward.is_active == True,
        db.or_(
            db.and_(EmployeeReward.is_recurring == False,
                    EmployeeReward.effective_date >= period_start,
                    EmployeeReward.effective_date <= period_end),
            EmployeeReward.is_recurring == True,
        ),
    ).scalar() or 0)
    total_deductions = float(db.session.query(db.func.coalesce(db.func.sum(EmployeeDeduction.amount), 0)).filter(
        EmployeeDeduction.user_id == salary.user_id,
        EmployeeDeduction.deduction_date >= period_start,
        EmployeeDeduction.deduction_date <= period_end,
    ).scalar() or 0)
    net = float(salary.base_salary) + total_rewards - total_deductions

    salary.status = 'paid'
    salary.paid_date = date.today()
    salary.approved_by = current_user.id
    salary.net_salary = net

    employee = salary.user if hasattr(salary, 'user') else None
    employee_name = getattr(employee, 'name', f'موظف #{salary.user_id}')
    expense = Expense(
        brand_id=salary.brand_id,
        branch_id=getattr(employee, 'branch_id', None),
        category_name='رواتب',
        amount=net,
        description=f'راتب {employee_name} - {salary.month_name} {salary.year}',
        date=salary.paid_date,
        status='approved',
        approved_by=current_user.id,
        approved_at=datetime.utcnow(),
        created_by=current_user.id,
    )
    db.session.add(expense)
    db.session.commit()
    flash(f'تم تأكيد دفع راتب {employee_name} وتسجيله كمصروف', 'success')
    return redirect(url_for('finance.salaries_list', month=salary.month, year=salary.year))


@finance_bp.route('/refunds')
@login_required
@finance_required
def refunds_list():
    """List refunds"""
    page, per_page = pagination_args(request)

    # Base query
    query = apply_branch_filter(Refund.query, Refund)

    # Pagination
    refunds = query.order_by(Refund.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return render_template('finance/refunds.html', refunds=refunds)


@finance_bp.route('/expenses/<int:expense_id>')
@login_required
@finance_required
def expense_view(expense_id):
    """View expense details"""
    expense = Expense.query.get_or_404(expense_id)

    if not current_user.can_access_brand(expense.brand_id):
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('finance.expenses'))

    return render_template('finance/expense_view.html', expense=expense)


# ─── GYM-32 — expense edit + soft-delete ──────────────────────────────────

@finance_bp.route('/expenses/<int:expense_id>/edit', methods=['GET', 'POST'])
@login_required
@finance_required
def expense_edit(expense_id):
    """Safe-field edit: amount, category, description, receipt. Status +
    branch + brand are intentionally fixed."""
    expense = Expense.query.get_or_404(expense_id)
    if not current_user.can_access_brand(expense.brand_id):
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('finance.expenses'))
    if expense.is_deleted:
        flash('المصروف محذوف.', 'warning')
        return redirect(url_for('finance.expenses'))

    categories = ExpenseCategory.query.filter_by(brand_id=expense.brand_id).all()

    if request.method == 'POST':
        try:
            new_amount = float(request.form.get('amount') or expense.amount)
        except ValueError:
            flash('قيمة المبلغ غير صالحة', 'danger')
            return redirect(url_for('finance.expense_edit', expense_id=expense.id))
        if new_amount <= 0:
            flash('المبلغ يجب أن يكون أكبر من صفر', 'danger')
            return redirect(url_for('finance.expense_edit', expense_id=expense.id))

        cat_name = (request.form.get('category_name') or expense.category_name).strip()
        cat_row = ExpenseCategory.query.filter_by(
            brand_id=expense.brand_id, name=cat_name
        ).first()

        expense.amount = new_amount
        expense.category_name = cat_name
        if cat_row:
            expense.category_id = cat_row.id
        expense.description = (request.form.get('description') or '').strip() or None

        # Optional new receipt
        receipt = request.files.get('receipt_image')
        if receipt and getattr(receipt, 'filename', ''):
            saved = save_uploaded_file(receipt, folder='receipts')
            if saved:
                expense.receipt_image = saved

        db.session.commit()
        flash('تم تحديث المصروف', 'success')
        return redirect(url_for('finance.expense_view', expense_id=expense.id))

    return render_template('finance/expense_edit.html',
                           expense=expense, categories=categories)


@finance_bp.route('/expenses/<int:expense_id>/delete', methods=['POST'])
@login_required
@finance_required
def expense_delete(expense_id):
    """Soft-delete an expense (status check + brand check)."""
    expense = Expense.query.get_or_404(expense_id)
    if not current_user.can_access_brand(expense.brand_id):
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('finance.expenses'))
    if expense.is_deleted:
        flash('المصروف محذوف مسبقاً.', 'warning')
        return redirect(url_for('finance.expenses'))

    expense.is_deleted = True
    db.session.commit()
    flash(f'تم حذف المصروف #{expense.id}.', 'success')
    return redirect(url_for('finance.expenses'))


@finance_bp.route('/expenses/<int:expense_id>/approve', methods=['POST'])
@login_required
@brand_manager_required
def expense_approve(expense_id):
    """Approve expense"""
    expense = Expense.query.get_or_404(expense_id)

    if not current_user.can_access_brand(expense.brand_id):
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('finance.expenses'))

    if expense.status != 'pending':
        flash('هذا المصروف تم معالجته مسبقاً', 'warning')
        return redirect(url_for('finance.expense_view', expense_id=expense_id))

    expense.status = 'approved'
    expense.approved_by = current_user.id
    expense.approved_at = datetime.utcnow()

    db.session.commit()
    flash('تم اعتماد المصروف', 'success')

    return redirect(url_for('finance.expenses', status='pending'))


@finance_bp.route('/expenses/<int:expense_id>/reject', methods=['GET', 'POST'])
@login_required
@brand_manager_required
def expense_reject(expense_id):
    """Reject expense"""
    expense = Expense.query.get_or_404(expense_id)

    if not current_user.can_access_brand(expense.brand_id):
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('finance.expenses'))

    if expense.status != 'pending':
        flash('هذا المصروف تم معالجته مسبقاً', 'warning')
        return redirect(url_for('finance.expense_view', expense_id=expense_id))

    if request.method == 'POST':
        reason = request.form.get('rejection_reason', '')
        expense.status = 'rejected'
        expense.approved_by = current_user.id
        expense.approved_at = datetime.utcnow()
        expense.rejection_reason = reason

        db.session.commit()
        flash('تم رفض المصروف', 'info')

        return redirect(url_for('finance.expenses', status='pending'))

    return render_template('finance/expense_reject.html', expense=expense)


# ============== xlsx exports — GYM-16 rework ==============

def _xlsx_response(filename, headers, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = filename[:30]
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0F3460')
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max((len(str(c.value or '')) for c in col), default=10) + 2, 40)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'{filename}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _date_range_args():
    df = request.args.get('date_from', date.today().replace(day=1).isoformat())
    dt = request.args.get('date_to', date.today().isoformat())
    try:
        return date.fromisoformat(df), date.fromisoformat(dt)
    except ValueError:
        return date.today().replace(day=1), date.today()


def _scoped_brand_id():
    """Same scoping the finance pages use — admin picks, everyone else pinned."""
    if current_user.can_view_all_brands:
        return request.args.get('brand_id', type=int)
    return current_user.brand_id


@finance_bp.route('/income/export.xlsx')
@login_required
@finance_required
def income_export():
    fd, td = _date_range_args()
    q = Income.query.filter(Income.date >= fd, Income.date <= td,
                            Income.is_deleted == False)  # GYM-32
    bid = _scoped_brand_id()
    if bid:
        q = q.filter(Income.brand_id == bid)
    if current_user.branch_id and not current_user.can_view_all_brands:
        q = q.filter(Income.branch_id == current_user.branch_id)
    pm = request.args.get('payment_method', '')
    if pm:
        q = q.filter(Income.payment_method == pm)
    rows = [[i.date.isoformat(), float(i.amount or 0),
             i.type or '', i.payment_method or '', i.description or '',
             i.branch.name if i.branch else '-']
            for i in q.order_by(Income.date.desc()).all()]
    return _xlsx_response(
        f'income-{fd.isoformat()}-to-{td.isoformat()}',
        ['التاريخ', 'المبلغ', 'النوع', 'طريقة الدفع', 'الوصف', 'الفرع'],
        rows,
    )


@finance_bp.route('/expenses/export.xlsx')
@login_required
@finance_required
def expenses_export():
    fd, td = _date_range_args()
    q = Expense.query.filter(Expense.date >= fd, Expense.date <= td)
    bid = _scoped_brand_id()
    if bid:
        q = q.filter(Expense.brand_id == bid)
    if current_user.branch_id and not current_user.can_view_all_brands:
        q = q.filter(Expense.branch_id == current_user.branch_id)
    st = request.args.get('status', '')
    if st:
        q = q.filter(Expense.status == st)
    cat = request.args.get('category', '')
    if cat:
        q = q.filter(Expense.category_name == cat)
    rows = [[e.date.isoformat(), float(e.amount or 0),
             e.category_name or '', e.status or '', e.description or '',
             e.creator.name if e.creator else '-']
            for e in q.order_by(Expense.date.desc()).all()]
    return _xlsx_response(
        f'expenses-{fd.isoformat()}-to-{td.isoformat()}',
        ['التاريخ', 'المبلغ', 'الفئة', 'الحالة', 'الوصف', 'بواسطة'],
        rows,
    )


@finance_bp.route('/sales-transactions/export.xlsx')
@login_required
@finance_required
def sales_transactions_export():
    fd, td = _date_range_args()
    q = db.session.query(SubscriptionPayment).join(Subscription).join(Subscription.member)
    bid = _scoped_brand_id()
    if bid:
        q = q.filter(Subscription.brand_id == bid)
    else:
        q = q.filter(Subscription.brand_id == current_user.brand_id)
    q = q.filter(db.func.date(SubscriptionPayment.payment_date) >= fd,
                 db.func.date(SubscriptionPayment.payment_date) <= td)
    pm = request.args.get('payment_method', '')
    if pm:
        q = q.filter(SubscriptionPayment.payment_method == pm)
    rows = []
    for p in q.order_by(SubscriptionPayment.payment_date.desc()).all():
        sub = p.subscription
        m = sub.member if sub else None
        rows.append([
            p.payment_date.strftime('%Y-%m-%d %H:%M') if p.payment_date else '',
            m.name if m else '-',
            sub.plan.name if sub and sub.plan else '-',
            float(p.amount or 0),
            p.payment_method or '',
            sub.branch.name if sub and sub.branch else '-',
        ])
    return _xlsx_response(
        f'sales-transactions-{fd.isoformat()}-to-{td.isoformat()}',
        ['التاريخ', 'العضو', 'الباقة', 'المبلغ', 'طريقة الدفع', 'الفرع'],
        rows,
    )


@finance_bp.route('/salaries/export.xlsx')
@login_required
@finance_required
def salaries_export():
    month = request.args.get('month', date.today().month, type=int)
    year = request.args.get('year', date.today().year, type=int)
    q = apply_branch_filter(Salary.query, Salary).filter_by(month=month, year=year)
    brand_name_by_id = {b.id: b.name for b in Brand.query.all()}
    rows = []
    for s in q.all():
        rows.append([
            s.user.name if s.user else '-',
            brand_name_by_id.get(s.brand_id, '-'),
            float(s.base_salary or 0),
            float(s.net_salary or 0),
            s.status_text,
            s.paid_date.isoformat() if s.paid_date else '',
        ])
    return _xlsx_response(
        f'salaries-{year}-{month:02d}',
        ['الموظف', 'البراند', 'الراتب الأساسي', 'الصافي', 'الحالة', 'تاريخ الدفع'],
        rows,
    )
