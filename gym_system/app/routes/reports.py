from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from datetime import date, timedelta
from sqlalchemy import func, case
import io

from app import db
from app.models.company import Brand, Branch
from app.models.user import User
from app.models.member import Member
from app.models.subscription import Subscription
from app.models.finance import Income, Expense
from app.models.service import ServiceType
from app.models.giftcard import GiftCard
from app.models.offer import PromotionalOffer

reports_bp = Blueprint('reports', __name__)


def _xlsx_response(filename, headers, rows):
    """Build an xlsx file in-memory and return a Flask response."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = filename[:30]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0F3460')
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'{filename}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@reports_bp.route('/')
@login_required
def index():
    """Reports index - redirect to staff performance or show menu"""
    if current_user.can_view_reports:
        return redirect(url_for('reports.staff_performance'))
    flash('ليس لديك صلاحية', 'danger')
    return redirect(url_for('dashboard.index'))


@reports_bp.route('/staff-performance')
@login_required
def staff_performance():
    """Staff performance report - shows revenue and subscriptions by employee"""
    if not current_user.can_view_reports:
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))

    # Brand scope: cross-brand viewers can pick any brand; everyone else is
    # pinned to their own brand_id (closes the GYM-16 leak — without this the
    # owner would get all-brand aggregates whenever ?brand_id= wasn't set).
    if current_user.can_view_all_brands:
        brand_id = request.args.get('brand_id', type=int)
    else:
        brand_id = current_user.brand_id
    branch_id = request.args.get('branch_id', type=int)
    period = request.args.get('period', 'month')  # week, month, quarter, year
    
    today = date.today()
    
    # Calculate date range based on period
    if period == 'week':
        start_date = today - timedelta(days=7)
    elif period == 'month':
        start_date = today.replace(day=1)
    elif period == 'quarter':
        quarter_month = ((today.month - 1) // 3) * 3 + 1
        start_date = today.replace(month=quarter_month, day=1)
    elif period == 'year':
        start_date = today.replace(month=1, day=1)
    else:
        start_date = today.replace(day=1)
    
    # Base query for employees
    emp_query = User.query.filter(User.is_active == True, User.is_deleted == False)  # GYM-43
    
    if brand_id:
        emp_query = emp_query.filter(User.brand_id == brand_id)
    if branch_id:
        emp_query = emp_query.filter(User.branch_id == branch_id)
    
    employees = emp_query.all()
    
    # Get performance data for each employee
    staff_data = []
    for emp in employees:
        # GYM-32 — exclude soft-deleted subscriptions from this employee's stats.
        subs_created = Subscription.query.filter(
            Subscription.created_by == emp.id,
            Subscription.created_at >= start_date,
            Subscription.is_deleted == False,
        ).count()

        revenue = db.session.query(func.sum(Subscription.total_amount)).filter(
            Subscription.created_by == emp.id,
            Subscription.created_at >= start_date,
            Subscription.is_deleted == False,
        ).scalar() or 0

        renewals = db.session.query(Subscription).filter(
            Subscription.created_by == emp.id,
            Subscription.created_at >= start_date,
            Subscription.is_deleted == False,
            Subscription.member_id.in_(
                db.session.query(Subscription.member_id).group_by(
                    Subscription.member_id
                ).having(func.count(Subscription.id) > 1)
            )
        ).count()
        
        # New members (first subscription)
        new_members = subs_created - renewals
        
        # Renewal rate
        renewal_rate = round((renewals / subs_created * 100) if subs_created > 0 else 0, 1)
        
        # Average subscription value
        avg_value = round(float(revenue) / subs_created if subs_created > 0 else 0, 0)
        
        staff_data.append({
            'id': emp.id,
            'name': emp.name,
            'role': emp.role.name if emp.role else '-',
            'brand': emp.brand.name if emp.brand else '-',
            'branch': emp.branch.name if emp.branch else '-',
            'subscriptions': subs_created,
            'new_members': new_members,
            'renewals': renewals,
            'renewal_rate': renewal_rate,
            'revenue': float(revenue),
            'avg_value': avg_value
        })
    
    # Sort by revenue (descending)
    staff_data = sorted(staff_data, key=lambda x: x['revenue'], reverse=True)
    
    # Add ranking
    for i, emp in enumerate(staff_data):
        emp['rank'] = i + 1
    
    # Get top and bottom performers
    top_performers = staff_data[:5] if len(staff_data) >= 5 else staff_data
    bottom_performers = staff_data[-5:][::-1] if len(staff_data) >= 5 else []
    
    # Brand selector: cross-brand viewers see every active brand; everyone
    # else gets just their own (so the dropdown doesn't leak brand names).
    if current_user.can_view_all_brands:
        brands = Brand.query.filter_by(is_active=True).all()
    elif current_user.brand_id:
        brands = Brand.query.filter_by(id=current_user.brand_id, is_active=True).all()
    else:
        brands = []
    branches = []
    if brand_id:
        branches = Branch.query.filter_by(brand_id=brand_id, is_active=True).all()
    
    # Calculate totals
    totals = {
        'subscriptions': sum(e['subscriptions'] for e in staff_data),
        'revenue': sum(e['revenue'] for e in staff_data),
        'renewals': sum(e['renewals'] for e in staff_data),
        'new_members': sum(e['new_members'] for e in staff_data)
    }
    
    return render_template('reports/staff_performance.html',
                          staff_data=staff_data,
                          top_performers=top_performers,
                          bottom_performers=bottom_performers,
                          brands=brands,
                          branches=branches,
                          totals=totals,
                          selected_brand=brand_id,
                          selected_branch=branch_id,
                          period=period,
                          start_date=start_date,
                          today=today)


@reports_bp.route('/financial')
@login_required
def financial():
    """Financial intelligence report - Gift cards, offers, payment methods analysis"""
    if not current_user.can_view_reports:
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))

    # See staff_performance — same GYM-16 leak fix.
    if current_user.can_view_all_brands:
        brand_id = request.args.get('brand_id', type=int)
    else:
        brand_id = current_user.brand_id
    period = request.args.get('period', 'month')
    
    today = date.today()
    
    # Calculate date range
    if period == 'week':
        start_date = today - timedelta(days=7)
    elif period == 'month':
        start_date = today.replace(day=1)
    elif period == 'quarter':
        quarter_month = ((today.month - 1) // 3) * 3 + 1
        start_date = today.replace(month=quarter_month, day=1)
    elif period == 'year':
        start_date = today.replace(month=1, day=1)
    else:
        start_date = today.replace(day=1)
    
    # Build base filter. GYM-32 — soft-deleted income + subscriptions stay
    # out of the financial report.
    income_filter = [Income.date >= start_date, Income.date <= today,
                     Income.is_deleted == False]
    sub_filter = [Subscription.created_at >= start_date,
                  Subscription.is_deleted == False]
    gc_filter = [func.date(GiftCard.created_at) >= start_date]
    
    if brand_id:
        income_filter.append(Income.brand_id == brand_id)
        sub_filter.append(Subscription.brand_id == brand_id)
        gc_filter.append(GiftCard.brand_id == brand_id)
    
    # === PAYMENT METHOD ANALYSIS ===
    payment_breakdown = db.session.query(
        Income.payment_method,
        func.sum(Income.amount).label('amount'),
        func.count(Income.id).label('transactions')
    ).filter(*income_filter).group_by(Income.payment_method).all()
    
    payment_stats = {
        'cash': {'amount': 0, 'transactions': 0},
        'card': {'amount': 0, 'transactions': 0},
        'transfer': {'amount': 0, 'transactions': 0}
    }
    total_income = 0
    for method, amount, transactions in payment_breakdown:
        if method in payment_stats:
            payment_stats[method] = {
                'amount': float(amount or 0),
                'transactions': transactions
            }
            total_income += float(amount or 0)
    
    # Calculate percentages
    for method in payment_stats:
        payment_stats[method]['percentage'] = round(
            (payment_stats[method]['amount'] / total_income * 100) if total_income > 0 else 0, 1
        )
    
    # === GIFT CARD ANALYTICS ===
    # Gift cards sold
    gc_sold_query = db.session.query(
        func.count(GiftCard.id).label('count'),
        func.sum(GiftCard.original_amount).label('total')
    ).filter(*gc_filter)
    gc_sold = gc_sold_query.first()
    
    # Gift cards redeemed
    gc_redeemed_filter = gc_filter + [GiftCard.status == 'redeemed']
    gc_redeemed = db.session.query(
        func.count(GiftCard.id).label('count'),
        func.sum(GiftCard.original_amount).label('total')
    ).filter(*gc_redeemed_filter).first()
    
    # Outstanding liability (active gift cards)
    if brand_id:
        outstanding = db.session.query(
            func.sum(GiftCard.remaining_amount)
        ).filter(
            GiftCard.brand_id == brand_id,
            GiftCard.status.in_(['active', 'partially_used'])
        ).scalar() or 0
    else:
        outstanding = db.session.query(
            func.sum(GiftCard.remaining_amount)
        ).filter(
            GiftCard.status.in_(['active', 'partially_used'])
        ).scalar() or 0
    
    gift_card_stats = {
        'sold_count': gc_sold.count or 0,
        'sold_amount': float(gc_sold.total or 0),
        'redeemed_count': gc_redeemed.count or 0,
        'redeemed_amount': float(gc_redeemed.total or 0),
        'outstanding': float(outstanding),
        'redemption_rate': round((gc_redeemed.count / gc_sold.count * 100) if gc_sold.count else 0, 1)
    }
    
    # === PROMOTIONAL VS NON-PROMOTIONAL REVENUE ===
    # Revenue with offers
    promo_revenue = db.session.query(
        func.sum(Subscription.total_amount).label('revenue'),
        func.count(Subscription.id).label('count'),
        func.sum(Subscription.offer_discount).label('discount')
    ).filter(
        *sub_filter,
        Subscription.offer_id.isnot(None)
    ).first()
    
    # Revenue without offers
    non_promo_revenue = db.session.query(
        func.sum(Subscription.total_amount).label('revenue'),
        func.count(Subscription.id).label('count')
    ).filter(
        *sub_filter,
        Subscription.offer_id.is_(None)
    ).first()
    
    promo_stats = {
        'with_offer': {
            'revenue': float(promo_revenue.revenue or 0),
            'count': promo_revenue.count or 0,
            'discount_given': float(promo_revenue.discount or 0)
        },
        'without_offer': {
            'revenue': float(non_promo_revenue.revenue or 0),
            'count': non_promo_revenue.count or 0
        }
    }
    
    total_subs = promo_stats['with_offer']['count'] + promo_stats['without_offer']['count']
    promo_stats['with_offer']['percentage'] = round(
        (promo_stats['with_offer']['count'] / total_subs * 100) if total_subs > 0 else 0, 1
    )
    promo_stats['without_offer']['percentage'] = round(
        (promo_stats['without_offer']['count'] / total_subs * 100) if total_subs > 0 else 0, 1
    )
    
    # === TOP OFFERS BY USAGE ===
    top_offers = db.session.query(
        PromotionalOffer.name,
        PromotionalOffer.discount_type,
        PromotionalOffer.discount_value,
        func.count(Subscription.id).label('usage_count'),
        func.sum(Subscription.total_amount).label('revenue'),
        func.sum(Subscription.offer_discount).label('discount_given')
    ).join(
        Subscription, Subscription.offer_id == PromotionalOffer.id
    ).filter(
        Subscription.created_at >= start_date
    ).group_by(
        PromotionalOffer.id, PromotionalOffer.name, 
        PromotionalOffer.discount_type, PromotionalOffer.discount_value
    ).order_by(func.count(Subscription.id).desc()).limit(10).all()
    
    offers_data = []
    for offer in top_offers:
        offers_data.append({
            'name': offer.name,
            'discount': f"{offer.discount_value}{'%' if offer.discount_type == 'percentage' else ' ر.س'}",
            'usage': offer.usage_count,
            'revenue': float(offer.revenue or 0),
            'discount_given': float(offer.discount_given or 0)
        })
    
    # === REVENUE BY SERVICE TYPE ===
    service_revenue = db.session.query(
        ServiceType.name,
        ServiceType.category,
        func.sum(Income.amount).label('revenue'),
        func.count(Income.id).label('transactions')
    ).join(
        Income, Income.service_type_id == ServiceType.id
    ).filter(*income_filter).group_by(
        ServiceType.id, ServiceType.name, ServiceType.category
    ).order_by(func.sum(Income.amount).desc()).all()
    
    service_data = []
    for svc in service_revenue:
        service_data.append({
            'name': svc.name,
            'category': svc.category,
            'revenue': float(svc.revenue or 0),
            'transactions': svc.transactions
        })
    
    # Brand selector — same scoping as staff_performance.
    if current_user.can_view_all_brands:
        brands = Brand.query.filter_by(is_active=True).all()
    elif current_user.brand_id:
        brands = Brand.query.filter_by(id=current_user.brand_id, is_active=True).all()
    else:
        brands = []

    return render_template('reports/financial.html',
                          payment_stats=payment_stats,
                          total_income=total_income,
                          gift_card_stats=gift_card_stats,
                          promo_stats=promo_stats,
                          offers_data=offers_data,
                          service_data=service_data,
                          brands=brands,
                          selected_brand=brand_id,
                          period=period,
                          start_date=start_date,
                          today=today)


@reports_bp.route('/attendance')
@login_required
def attendance():
    """Attendance report — totals + daily average per brand."""
    if not current_user.can_view_reports:
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))

    from app.models.attendance import MemberAttendance

    period = request.args.get('period', 'month')
    today = date.today()

    if period == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif period == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today
    elif period == 'custom':
        try:
            start_date = date.fromisoformat(request.args.get('start_date') or '')
        except (TypeError, ValueError):
            start_date = today.replace(day=1)
        try:
            end_date = date.fromisoformat(request.args.get('end_date') or '')
        except (TypeError, ValueError):
            end_date = today
    else:  # month (default)
        start_date = today.replace(day=1)
        end_date = today

    brand_id = request.args.get('brand_id', type=int)

    # Brands the user can see
    if current_user.is_owner or current_user.can_view_all_brands:
        brand_query = Brand.query.filter_by(is_active=True)
    else:
        brand_query = Brand.query.filter_by(id=current_user.brand_id, is_active=True)
    if brand_id:
        brand_query = brand_query.filter_by(id=brand_id)
    brands_in_scope = brand_query.all()

    days_count = max((end_date - start_date).days + 1, 1)
    report_data = []
    for b in brands_in_scope:
        total = MemberAttendance.query.filter(
            MemberAttendance.brand_id == b.id,
            func.date(MemberAttendance.check_in) >= start_date,
            func.date(MemberAttendance.check_in) <= end_date,
        ).count()
        report_data.append({
            'brand':            b,
            'total_attendance': total,
            'daily_average':    total / days_count,
            'days':             days_count,
        })

    # Filter dropdown source — all brands the user is allowed to see
    brands = Brand.query.filter_by(is_active=True).all() if current_user.is_owner else []

    return render_template('reports/attendance.html',
                           report_data=report_data,
                           brands=brands,
                           period=period,
                           start_date=start_date.isoformat(),
                           end_date=end_date.isoformat())


# ============ GYM-16 Excel exports ============

@reports_bp.route('/staff-performance/export.xlsx')
@login_required
def staff_performance_export():
    """xlsx export of /reports/staff-performance with the SAME brand scoping."""
    if not current_user.can_view_reports:
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))

    if current_user.can_view_all_brands:
        brand_id = request.args.get('brand_id', type=int)
    else:
        brand_id = current_user.brand_id
    branch_id = request.args.get('branch_id', type=int)
    period = request.args.get('period', 'month')
    today = date.today()
    if period == 'week':
        start_date = today - timedelta(days=7)
    elif period == 'quarter':
        qm = ((today.month - 1) // 3) * 3 + 1
        start_date = today.replace(month=qm, day=1)
    elif period == 'year':
        start_date = today.replace(month=1, day=1)
    else:
        start_date = today.replace(day=1)

    emp_query = User.query.filter(User.is_active == True, User.is_deleted == False)  # GYM-43
    if brand_id:
        emp_query = emp_query.filter(User.brand_id == brand_id)
    if branch_id:
        emp_query = emp_query.filter(User.branch_id == branch_id)

    rows = []
    for emp in emp_query.all():
        # GYM-40 — exclude soft-deleted subscriptions from per-staff totals.
        subs = Subscription.query.filter(
            Subscription.created_by == emp.id,
            Subscription.created_at >= start_date,
            Subscription.is_deleted == False,
        ).count()
        revenue = db.session.query(func.coalesce(func.sum(Subscription.total_amount), 0)).filter(
            Subscription.created_by == emp.id,
            Subscription.created_at >= start_date,
            Subscription.is_deleted == False,
        ).scalar() or 0
        rows.append([
            emp.name, emp.role.name if emp.role else '-',
            emp.brand.name if emp.brand else '-',
            emp.branch.name if emp.branch else '-',
            subs, float(revenue),
        ])
    rows.sort(key=lambda r: r[5], reverse=True)
    return _xlsx_response(
        f'staff-performance-{start_date.isoformat()}-to-{today.isoformat()}',
        ['الموظف', 'الوظيفة', 'البراند', 'الفرع', 'الاشتراكات', 'الإيرادات'],
        rows,
    )


@reports_bp.route('/financial/export.xlsx')
@login_required
def financial_export():
    """xlsx export of /reports/financial — payment-method breakdown."""
    if not current_user.can_view_reports:
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))

    if current_user.can_view_all_brands:
        brand_id = request.args.get('brand_id', type=int)
    else:
        brand_id = current_user.brand_id
    period = request.args.get('period', 'month')
    today = date.today()
    if period == 'week':
        start_date = today - timedelta(days=7)
    elif period == 'quarter':
        qm = ((today.month - 1) // 3) * 3 + 1
        start_date = today.replace(month=qm, day=1)
    elif period == 'year':
        start_date = today.replace(month=1, day=1)
    else:
        start_date = today.replace(day=1)

    flt = [Income.date >= start_date, Income.date <= today,
           Income.is_deleted == False]  # GYM-32
    if brand_id:
        flt.append(Income.brand_id == brand_id)
    payment_breakdown = db.session.query(
        Income.payment_method,
        func.sum(Income.amount).label('amount'),
        func.count(Income.id).label('transactions')
    ).filter(*flt).group_by(Income.payment_method).all()

    rows = [[(method or '-'), int(transactions or 0), float(amount or 0)]
            for method, amount, transactions in payment_breakdown]
    return _xlsx_response(
        f'financial-{start_date.isoformat()}-to-{today.isoformat()}',
        ['طريقة الدفع', 'عدد العمليات', 'الإجمالي'],
        rows,
    )


@reports_bp.route('/daily-closing/export.xlsx')
@login_required
def daily_closing_export():
    """xlsx export of the GYM-17 daily-closing detailed report."""
    from app.models.daily_closing import DailyClosing
    from app.utils.helpers import apply_branch_filter, resolve_owner_branch_filter
    if not (current_user.role and current_user.role.can_view_daily_closing):
        flash('ليس لديك صلاحية', 'danger')
        return redirect(url_for('dashboard.index'))

    today = date.today()
    start_str = request.args.get('start_date', (today - timedelta(days=30)).isoformat())
    end_str = request.args.get('end_date', today.isoformat())
    try:
        start_date = date.fromisoformat(start_str)
        end_date = date.fromisoformat(end_str)
    except ValueError:
        start_date, end_date = today - timedelta(days=30), today

    base = apply_branch_filter(DailyClosing.query, DailyClosing,
                               branch_filter_id=resolve_owner_branch_filter())
    closings = base.filter(
        DailyClosing.closing_date >= start_date,
        DailyClosing.closing_date <= end_date,
    ).order_by(DailyClosing.closing_date.desc()).all()

    rows = []
    for c in closings:
        rows.append([
            c.closing_date.isoformat(),
            c.branch.name if c.branch else 'بدون فرع',
            float(c.total_sales or 0),
            float(c.cash_amount or 0),
            float(c.card_amount or 0),
            float(c.transfer_amount or 0),
            float(c.total_expenses or 0),
            float(c.cash_difference or 0),
            c.status,
        ])
    return _xlsx_response(
        f'daily-closing-{start_date.isoformat()}-to-{end_date.isoformat()}',
        ['التاريخ', 'الفرع', 'المبيعات', 'نقدي', 'شبكة', 'حوالة',
         'المصروفات', 'الفرق', 'الحالة'],
        rows,
    )
